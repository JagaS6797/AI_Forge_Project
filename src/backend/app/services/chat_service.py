from __future__ import annotations

import base64
import json
import logging
from io import BytesIO
from typing import AsyncIterator

from fastapi import HTTPException
from langchain_core.messages import HumanMessage, SystemMessage
from openpyxl import load_workbook
from openai import OpenAIError
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.ai.llm import chat_llm
from app.core.config import settings
from app.models.chat_message import ChatMessage
from app.models.file_attachment import FileAttachment
from app.models.user import User
from app.schemas.chat import ChatHistoryItem
from app.services.attachment_service import get_file_content
from app.services.image_generation_service import (
    extract_image_prompt,
    generate_image_attachment,
)
from app.services.rag_service import answer_with_rag, thread_has_rag_documents
from app.services.user_service import get_or_create_user_id


logger = logging.getLogger(__name__)


async def get_attachments_by_ids(
    db: AsyncSession,
    attachment_ids: list[str] | None,
) -> list[FileAttachment]:
    """Fetch attachment metadata by IDs."""
    if not attachment_ids:
        return []
    
    try:
        attachments = await db.scalars(
            select(FileAttachment).where(FileAttachment.id.in_(attachment_ids))
        )
        return list(attachments.all())
    except Exception:
        logger.exception("Failed to fetch attachments")
        return []


def _format_history(history: list[ChatHistoryItem]) -> str:
    if not history:
        return "(none)"

    return "\n".join(f"{item.role}: {item.content}" for item in history)


def _format_excel_content(file_name: str, content: bytes) -> str:
    try:
        workbook = load_workbook(filename=BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:
        logger.warning("Could not parse Excel file %s: %s", file_name, exc)
        return f"📊 EXCEL FILE: {file_name} (could not parse workbook content)"

    lines: list[str] = [f"📊 EXCEL FILE: {file_name}"]
    lines.append("Sheets: " + ", ".join(workbook.sheetnames[:10]))

    max_sheets = 3
    max_rows_per_sheet = 20
    max_cols_per_row = 12

    for sheet in workbook.worksheets[:max_sheets]:
        lines.append(f"\nSheet: {sheet.title}")
        row_counter = 0
        for row in sheet.iter_rows(values_only=True):
            if row_counter >= max_rows_per_sheet:
                lines.append("... [sheet rows truncated]")
                break

            cells = []
            for cell in row[:max_cols_per_row]:
                if cell is None:
                    cells.append("")
                else:
                    text = str(cell).replace("\n", " ").strip()
                    cells.append(text)

            if any(cells):
                lines.append(" | ".join(cells))
                row_counter += 1

        if row_counter == 0:
            lines.append("(no readable rows found)")

    workbook.close()
    return "\n".join(lines)


async def _format_attachments(
    db: AsyncSession,
    attachment_ids: list[str] | None,
) -> str:
    """Format attachment content for LLM context. Returns formatted string or empty string."""
    if not attachment_ids:
        return ""
    
    try:
        attachments = await db.scalars(
            select(FileAttachment).where(FileAttachment.id.in_(attachment_ids))
        )
        attachment_list = list(attachments.all())
        
        if not attachment_list:
            return ""
        
        formatted_parts = ["ATTACHMENTS PROVIDED:"]
        
        for att in attachment_list:
            # Categorize by MIME type and format appropriately
            if att.file_type.startswith("image/"):
                # For images, provide description to user can download
                formatted_parts.append(f"📷 IMAGE: {att.file_name}")
                formatted_parts.append(f"   Type: {att.file_type}")
                formatted_parts.append(f"   Size: {att.file_size / 1024:.1f} KB")
                formatted_parts.append(f"   [Available for download - user can view and provide details]")
                
            elif att.file_type.startswith("video/"):
                formatted_parts.append(f"🎬 VIDEO: {att.file_name}")
                formatted_parts.append(f"   Type: {att.file_type}")
                formatted_parts.append(f"   Size: {att.file_size / (1024*1024):.1f} MB")
                formatted_parts.append(f"   [Available for download]")
                
            elif att.file_type in ("text/plain", "text/markdown", "text/csv") or att.file_type.startswith("text/"):
                # For text files, include full content
                content = get_file_content(att.file_path)
                if content:
                    try:
                        text = content.decode("utf-8")
                        # Limit content to first 4000 chars to avoid token explosion
                        if len(text) > 4000:
                            text = text[:4000] + "\n\n... [content truncated]"
                        formatted_parts.append(f"📄 TEXT FILE: {att.file_name}")
                        formatted_parts.append(f"Content:\n{text}")
                    except Exception as e:
                        logger.warning(f"Could not decode text file {att.file_name}: {e}")
                        formatted_parts.append(f"📄 TEXT FILE: {att.file_name} (could not decode - may be binary)")
                else:
                    formatted_parts.append(f"📄 TEXT FILE: {att.file_name} (could not read from disk)")
                    
            elif att.file_type == "application/pdf":
                formatted_parts.append(f"📕 PDF DOCUMENT: {att.file_name}")
                formatted_parts.append(f"   Size: {att.file_size / (1024*1024):.1f} MB")
                formatted_parts.append(f"   [Available for download]")

            elif (
                att.file_type == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                or att.file_name.lower().endswith(".xlsx")
            ):
                content = get_file_content(att.file_path)
                if content:
                    formatted_parts.append(_format_excel_content(att.file_name, content))
                else:
                    formatted_parts.append(f"📊 EXCEL FILE: {att.file_name} (could not read from disk)")
                
            elif "code" in att.file_type or any(ext in att.file_name.lower() for ext in [".py", ".js", ".ts", ".java", ".cpp", ".go"]):
                # Treat as code file
                content = get_file_content(att.file_path)
                if content:
                    try:
                        text = content.decode("utf-8")
                        if len(text) > 4000:
                            text = text[:4000] + "\n\n... [code truncated]"
                        formatted_parts.append(f"💻 CODE FILE: {att.file_name}")
                        formatted_parts.append(f"```\n{text}\n```")
                    except Exception:
                        formatted_parts.append(f"💻 CODE FILE: {att.file_name} (could not decode)")
                else:
                    formatted_parts.append(f"💻 CODE FILE: {att.file_name} (could not read)")
                    
            else:
                formatted_parts.append(f"📦 FILE: {att.file_name}")
                formatted_parts.append(f"   Type: {att.file_type}")
                formatted_parts.append(f"   Size: {att.file_size / 1024:.1f} KB")
        
        return "\n".join(formatted_parts)
        
    except Exception as e:
        logger.exception(f"Failed to format attachments for LLM context: {e}")
        return "Attachments provided but could not be processed."


def _recent_conversations(
    history: list[ChatHistoryItem],
    *,
    max_conversations: int,
) -> list[ChatHistoryItem]:
    if max_conversations <= 0 or not history:
        return []

    user_count = 0
    start_index = len(history)

    for index in range(len(history) - 1, -1, -1):
        if history[index].role == "user":
            user_count += 1
            if user_count == max_conversations:
                start_index = index
                break

    if user_count < max_conversations:
        return history

    return history[start_index:]


async def save_chat_message(
    *,
    db: AsyncSession,
    user_email: str,
    role: str,
    content: str,
    thread_id: str | None = None,
    attachment_ids: list[str] | None = None,
) -> ChatMessage:
    user_id = await get_or_create_user_id(db, user_email)
    message = ChatMessage(
        user_id=user_id,
        role=role,
        content=content,
        thread_id=thread_id,
        attachment_ids=attachment_ids or [],
    )
    db.add(message)
    try:
        await db.commit()
    except Exception:
        await db.rollback()
        raise
    await db.refresh(message)
    
    # Link attachments to message by setting message_id
    if attachment_ids:
        try:
            attachments = await db.scalars(
                select(FileAttachment).where(FileAttachment.id.in_(attachment_ids))
            )
            for att in attachments.all():
                att.message_id = message.id
            await db.commit()
        except Exception:
            await db.rollback()
            logger.exception("Failed to link attachments to message")
    
    return message


async def list_chat_messages(
    *, db: AsyncSession, user_email: str, thread_id: str | None = None
) -> list[ChatMessage]:
    try:
        user = await db.scalar(select(User).where(User.email == user_email.lower().strip()))
    except Exception:
        logger.exception("Failed to load chat history from database.")
        return []

    if not user:
        return []

    try:
        query = (
            select(ChatMessage)
            .where(ChatMessage.user_id == user.id)
            .order_by(ChatMessage.created_at.asc())
        )
        if thread_id is not None:
            query = query.where(ChatMessage.thread_id == thread_id)
        result = await db.scalars(query)
        return list(result.all())
    except Exception:
        logger.exception("Failed to query chat messages.")
        return []


async def stream_chat_events(
    *,
    db: AsyncSession,
    message: str,
    thread_id: str,
    user_email: str,
    attachment_ids: list[str] | None = None,
    rag_enabled: bool = False,
) -> AsyncIterator[str]:
    from app.services.thread_service import maybe_set_thread_name

    # Validate that either message or attachments are provided
    message_text = message.strip() if message else ""
    has_attachments = bool(attachment_ids)
    
    if not message_text and not has_attachments:
        raise HTTPException(
            status_code=400,
            detail="Either a message or attachments must be provided"
        )
    assistant_chunks: list[str] = []

    try:
        # 1. Load existing history for this thread (before this new message)
        existing = await list_chat_messages(db=db, user_email=user_email, thread_id=thread_id)
        history_items = [ChatHistoryItem(role=m.role, content=m.content) for m in existing]
        memory_items = _recent_conversations(
            history_items,
            max_conversations=settings.chat_memory_conversations,
        )

        # 2. Format attachments and build multimodal content for LLM
        attachment_context = await _format_attachments(db, attachment_ids)
        attachment_records = await get_attachments_by_ids(db, attachment_ids)

        # 3. Auto-name thread on first message
        if not existing:
            new_name = await maybe_set_thread_name(db=db, thread_id=thread_id, first_message=message_text or "Attachments uploaded")
            if new_name:
                yield f"data: {json.dumps({'thread_name': new_name})}\n\n"

        # 4. Save user message with attachments
        try:
            await save_chat_message(
                db=db,
                user_email=user_email,
                role="user",
                content=message_text,
                thread_id=thread_id,
                attachment_ids=attachment_ids or [],
            )
        except Exception:
            await db.rollback()
            logger.exception("Failed to persist user message. Continuing chat stream.")

        image_prompt = extract_image_prompt(message_text)
        if image_prompt:
            generated_attachment, attachment_payload = await generate_image_attachment(
                db=db,
                user_email=user_email,
                prompt=image_prompt,
            )

            assistant_content = (
                "Image generated successfully. "
                "Use the preview below to open or download it."
            )
            yield f"data: {json.dumps({'token': assistant_content})}\n\n"
            yield f"data: {json.dumps({'attachment': attachment_payload})}\n\n"

            try:
                await save_chat_message(
                    db=db,
                    user_email=user_email,
                    role="assistant",
                    content=assistant_content,
                    thread_id=thread_id,
                    attachment_ids=[generated_attachment.id],
                )
            except Exception:
                await db.rollback()
                logger.exception("Failed to persist assistant image response. Continuing chat stream.")

            yield "data: {\"done\": true}\n\n"
            return

        if rag_enabled and message_text and await thread_has_rag_documents(db=db, user_email=user_email, thread_id=thread_id):
            rag_answer, should_use_rag = await answer_with_rag(
                db=db,
                user_email=user_email,
                thread_id=thread_id,
                history=_format_history(memory_items),
                question=message_text,
            )
            if should_use_rag and rag_answer:
                yield f"data: {json.dumps({'token': rag_answer})}\n\n"
                try:
                    await save_chat_message(
                        db=db,
                        user_email=user_email,
                        role="assistant",
                        content=rag_answer,
                        thread_id=thread_id,
                    )
                except Exception:
                    await db.rollback()
                    logger.exception("Failed to persist assistant RAG message. Continuing chat stream.")

                yield "data: {\"done\": true}\n\n"
                return
            elif not should_use_rag:
                yield "data: {\"event\": \"rag_fallback\"}\n\n"

        rag_documents_available = await thread_has_rag_documents(
            db=db,
            user_email=user_email,
            thread_id=thread_id,
        )

        if rag_documents_available and not rag_enabled:
            history_text = "(omitted because RAG is disabled for this thread)"
        else:
            history_text = _format_history(memory_items) if not has_attachments else "(omitted for attachment-focused query)"
        attachment_names = [att.file_name for att in attachment_records]

        user_content_blocks: list[dict[str, object]] = [
            {
                "type": "text",
                "text": (
                    "Conversation history:\n"
                    f"{history_text}\n\n"
                    "Attachment summary:\n"
                    f"{attachment_context or '(none)'}\n\n"
                    "Files attached in this request:\n"
                    f"{', '.join(attachment_names) if attachment_names else '(none)'}\n\n"
                    "User message:\n"
                    f"{message_text or '(attachment-only request)'}"
                ),
            }
        ]

        image_names_without_bytes: list[str] = []
        for att in attachment_records:
            if not att.file_type.startswith("image/"):
                continue

            image_bytes = get_file_content(att.file_path)
            if not image_bytes:
                image_names_without_bytes.append(att.file_name)
                continue

            image_b64 = base64.b64encode(image_bytes).decode("ascii")
            data_url = f"data:{att.file_type};base64,{image_b64}"
            user_content_blocks.append(
                {"type": "text", "text": f"Attached image file: {att.file_name}"}
            )
            user_content_blocks.append(
                {
                    "type": "image_url",
                    "image_url": {"url": data_url},
                }
            )

        if image_names_without_bytes:
            user_content_blocks.append(
                {
                    "type": "text",
                    "text": (
                        "Warning: The backend could not load bytes for these image files: "
                        + ", ".join(image_names_without_bytes)
                        + ". If analysis quality is limited, ask the user to re-upload these specific files."
                    ),
                }
            )

        system_prompt = (
            "You are a concise, helpful assistant for a web chatbot. "
            "When attachments are present, focus only on the files attached in THIS request unless the user explicitly asks to compare with earlier files. "
            "If image content blocks are present, analyze those images directly and reference the specific attached filename(s). "
            "Do not claim you cannot view an image when image content blocks are present. "
            "For videos, describe only what you can infer from provided metadata unless actual visual frames are provided."
        )

        if rag_documents_available and not rag_enabled:
            system_prompt += (
                " RAG is disabled for this thread. Do not use or infer anything from uploaded PDFs or prior PDF-derived answers. "
                "If the user asks about the uploaded PDF, attached document, thread document, or requests facts that depend on those documents, respond exactly with 'I do not know.'. "
                "Otherwise, answer using general knowledge only."
            )

        async for chunk in chat_llm.astream(
            [
                SystemMessage(content=system_prompt),
                HumanMessage(content=user_content_blocks),
            ],
            config={
                "metadata": {
                    "user_email": user_email,
                }
            },
        ):
            token = ""
            if isinstance(chunk.content, str):
                token = chunk.content
            elif isinstance(chunk.content, list):
                token = "".join(
                    part.get("text", "")
                    for part in chunk.content
                    if isinstance(part, dict) and part.get("type") == "text"
                )

            if token:
                assistant_chunks.append(token)
                yield f"data: {json.dumps({'token': token})}\n\n"

        assistant_content = "".join(assistant_chunks).strip()
        if assistant_content:
            try:
                await save_chat_message(
                    db=db,
                    user_email=user_email,
                    role="assistant",
                    content=assistant_content,
                    thread_id=thread_id,
                )
            except Exception:
                await db.rollback()
                logger.exception("Failed to persist assistant message. Continuing chat stream.")


        yield "data: {\"done\": true}\n\n"
    except OpenAIError as exc:
        logger.exception("Image/text model request failed")
        yield f"data: {json.dumps({'token': f'Image generation failed: {str(exc)}'})}\n\n"
        yield "data: {\"done\": true}\n\n"
    except HTTPException as exc:
        logger.exception("Chat request failed with HTTPException during stream")
        detail = exc.detail
        if isinstance(detail, dict):
            message = str(detail.get("message") or detail.get("error") or "Request failed")
        else:
            message = str(detail)
        yield f"data: {json.dumps({'token': message})}\n\n"
        yield "data: {\"done\": true}\n\n"
    except Exception as exc:
        logger.exception("Unexpected chat stream error")
        yield f"data: {json.dumps({'token': f'Unexpected error: {str(exc)}'})}\n\n"
        yield "data: {\"done\": true}\n\n"
